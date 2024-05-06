import configparser
import os
from openpyxl import load_workbook
import pandas as pd
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill

import algorithm
import setup


#############################################
########## Set Up Work Environment ##########
#############################################

App = setup.App()

workspace_directory = App.workspace_folder

config_ini = os.path.join(workspace_directory, "config.ini")

config = configparser.ConfigParser()
config.read(config_ini)

text_font = "Verdana"


######################################################
########## CARB Facility Matching Functions ##########
######################################################


def open_file_dialog():
    """
    This function allows the user to select an input excel file for the tool.
    It is called upon when the user clicks the GUi button associated with the
    open_button varaible in this script.
    """
    
    file_path = filedialog.askopenfilename()
    if file_path:
        file_path_entry.config(state='normal')
        file_path_entry.delete(0, tk.END)
        file_path_entry.insert(0, file_path)
        file_path_entry.config(state='readonly')

        excel_file_path = file_path_entry.get()
        workbook = load_workbook(excel_file_path)
        sheet_names = workbook.sheetnames

        selected_option.set('')
        dropdown['menu'].delete(0, 'end')

        new_choices = sheet_names

        for choice in new_choices:
            dropdown['menu'].add_command(label=choice, command=tk._setit(selected_option, choice))

def select_output_file():
    """
    This function allows the user to select an output excel file for the tool.
    It is called upon when the user clicks the GUI button associated with the
    open_button_output varaible in this script.
    """
    
    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if output_file_path:
        file_path_output.config(state='normal')
        file_path_output.delete(0, tk.END)
        file_path_output.insert(0, output_file_path)
        file_path_output.config(state='readonly')

def load_button_actions():
    """
    This function executes other functions (load_excel_file, add_field_matching,
    and add_output_file_selection) that all together display field configuation
    and output file selction within the GUI once an input file has been selected
    and loaded. It is called upon when the user clicks the GUI button associated
    with the load_button varaible in this script.
    """
    
    load_excel_file()
    add_field_matching()
    add_output_file_selection()

def load_excel_file():
    """
    This function loads in field names from the input excel file and uses these
    as selectable options within the dropdown bars that allow users to configure
    field names within the GUI.
    """
    excel_path = file_path_entry.get()
    sheet_name = selected_option.get()

    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    selected_option_CO.set('')
    dropdown_CO['menu'].delete(0, 'end')

    selected_option_AB.set('')
    dropdown_AB['menu'].delete(0, 'end')

    selected_option_DIS.set('')
    dropdown_DIS['menu'].delete(0, 'end')

    selected_option_FACID.set('')
    dropdown_FACID['menu'].delete(0, 'end')

    selected_option_FNAME.set('')
    dropdown_FNAME['menu'].delete(0, 'end')

    selected_option_FSTREET.set('')
    dropdown_FSTREET['menu'].delete(0, 'end')

    selected_option_FCITY.set('')
    dropdown_FCITY['menu'].delete(0, 'end')

    selected_option_FZIP.set('')
    dropdown_FZIP['menu'].delete(0, 'end')

    selected_option_FSIC.set('')
    dropdown_FSIC['menu'].delete(0, 'end')

    selected_option_FNAICS.set('')
    dropdown_FNAICS['menu'].delete(0, 'end')

    selected_option_Latitude.set('')
    dropdown_Latitude['menu'].delete(0, 'end')

    selected_option_Longitude.set('')
    dropdown_Longitude['menu'].delete(0, 'end')
    
    new_choices = df.columns

    for choice in new_choices:
        dropdown_CO['menu'].add_command(label=choice, command=tk._setit(selected_option_CO, choice))
        dropdown_AB['menu'].add_command(label=choice, command=tk._setit(selected_option_AB, choice))
        dropdown_DIS['menu'].add_command(label=choice, command=tk._setit(selected_option_DIS, choice))
        dropdown_FACID['menu'].add_command(label=choice, command=tk._setit(selected_option_FACID, choice))
        dropdown_FNAME['menu'].add_command(label=choice, command=tk._setit(selected_option_FNAME, choice))
        dropdown_FSTREET['menu'].add_command(label=choice, command=tk._setit(selected_option_FSTREET, choice))
        dropdown_FCITY['menu'].add_command(label=choice, command=tk._setit(selected_option_FCITY, choice))
        dropdown_FZIP['menu'].add_command(label=choice, command=tk._setit(selected_option_FZIP, choice))
        dropdown_FSIC['menu'].add_command(label=choice, command=tk._setit(selected_option_FSIC, choice))
        dropdown_FNAICS['menu'].add_command(label=choice, command=tk._setit(selected_option_FNAICS, choice))
        dropdown_Latitude['menu'].add_command(label=choice, command=tk._setit(selected_option_Latitude, choice))
        dropdown_Longitude['menu'].add_command(label=choice, command=tk._setit(selected_option_Longitude, choice))

def add_field_matching():
    """
    This function displays selected field names within the dropdowns of the
    field name configuration section of the GUI.
    """
    selection_label.grid(row=0, column=0, columnspan=2, pady=5, sticky='nsew')
    
    label_CO.grid(row=1, column=0, padx=5, sticky='w')
    dropdown_CO.grid(row=1, column=1)

    label_AB.grid(row=2, column=0, padx=5, sticky='w')
    dropdown_AB.grid(row=2, column=1)

    label_DIS.grid(row=3, column=0, padx=5, sticky='w')
    dropdown_DIS.grid(row=3, column=1)

    label_FACID.grid(row=4, column=0, padx=5, sticky='w')
    dropdown_FACID.grid(row=4, column=1)

    label_FNAME.grid(row=5, column=0, padx=5, sticky='w')
    dropdown_FNAME.grid(row=5, column=1)

    label_FSTREET.grid(row=6, column=0, padx=5, sticky='w')
    dropdown_FSTREET.grid(row=6, column=1)

    label_FCITY.grid(row=7, column=0, padx=5, sticky='w')
    dropdown_FCITY.grid(row=7, column=1)

    label_FZIP.grid(row=8, column=0, padx=5, sticky='w')
    dropdown_FZIP.grid(row=8, column=1)

    label_FSIC.grid(row=9, column=0, padx=5, sticky='w')
    dropdown_FSIC.grid(row=9, column=1)

    label_FNAICS.grid(row=10, column=0, padx=5, sticky='w')
    dropdown_FNAICS.grid(row=10, column=1)

    label_Latitude.grid(row=11, column=0, padx=5, sticky='w')
    dropdown_Latitude.grid(row=11, column=1)

    label_Longitude.grid(row=12, column=0, padx=5, sticky='w')
    dropdown_Longitude.grid(row=12, column=1)

    

def add_output_file_selection():
    """
    This function adds the selected output file location text to the GUI
    application.
    """
    label_output.grid(row=0, column=0, sticky='w')
    open_button_output.grid(row=0, column=1, padx=5)
    file_path_output.grid(row=0, column=2, sticky='e')

    execute_button.grid(row=13, column=0, columnspan=3, pady=10, sticky='nsew')

def execute_facility_matching():
    """
    This function calls upon and executes all of the functions within
    algorithm.py to run the facility matching process. Each section of this
    function is seperated by a print statement that notifies the users which
    stage of the process is being executed. Documentation for each algorthm.py
    function is included in that file. t is called upon when the user clicks the
    GUI button associated with the execute_button varaible in this script.
    """
    
    print("\n\n #### Starting Facility Matching #####\n")
    
    excel_path = file_path_entry.get()
    sheet_name = selected_option.get()
    
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.insert(0, 'UID', df.index)
    #df['UID'] = df.index

    print("Standardizing Input Table...")
    
    df_standardized = algorithm.standardize_table(
        df = df,
        CO_name = selected_option_CO.get(),
        AB_name = selected_option_AB.get(),
        DIS_name = selected_option_DIS.get(),
        FACID_name = selected_option_FACID.get(),
        FNAME_name = selected_option_FNAME.get(),
        FSTREET_name = selected_option_FSTREET.get(),
        FCITY_name = selected_option_FCITY.get(),
        FZIP_name = selected_option_FZIP.get(),
        FSIC_name = selected_option_FSIC.get(),
        FNAICS_name = selected_option_FNAICS.get(),
        LAT_name = selected_option_Latitude.get(),
        LON_name = selected_option_Longitude.get()
        )

    print("Reading in Master Facilities Table...")

    df_master = algorithm.read_in_master_table(
        db_loc = App.facility_database_file,
        table_name = config["Database"].get("Master_Facilities_Table_Name"),
        ARBID_name = config["Database"].get("AB_name"),
        CO_name = config["Database"].get("CO_name"),
        AB_name = config["Database"].get("AB_name"),
        DIS_name = config["Database"].get("DIS_name"),
        FACID_name = config["Database"].get("FACID_name"),
        FNAME_name = config["Database"].get("FNAME_name"),
        FSTREET_name = config["Database"].get("FSTREET_name"),
        FCITY_name = config["Database"].get("FCITY_name"),
        FZIP_name = config["Database"].get("FZIP_name"),
        FSIC_name = config["Database"].get("FSIC_name"),
        FNAICS_name = config["Database"].get("FNAICS_name"),
        LAT_name = config["Database"].get("LAT_name"),
        LON_name = config["Database"].get("LON_name"),
        PARCEL_name = config["Database"].get("PARCEL_name")
        )

    df_master.insert(0, 'UID', df_master.index)

    print("Loading Parcel Dataset...")

    parcel_gdf = algorithm.load_parcel_dataset(
        pqt_folder_path = App.parcel_parquet_folder
        )

    print("Performing Spatial Join with Parcel Dataset...")

    df_standardized = algorithm.run_spatial_join(df_standardized, parcel_gdf)
    df_master = algorithm.run_spatial_join(df_master, parcel_gdf)

    print("Standardizing Facility Name and Address Fields...")

    df_standardized = algorithm.standardize_text_fields(df_standardized, logic_path = os.path.join(workspace_directory, r"dev\standardization\Word_Replacement_Table.csv"))
    df_master = algorithm.standardize_text_fields(df_master, logic_path = os.path.join(workspace_directory, r"dev\standardization\Word_Replacement_Table.csv"))

    print("Rounding Coordinates to 5 Decimal Places...")

    df_standardized["LATITUDE_ROUND_5"] = df_standardized["LAT_NAD83"].round(5)
    df_standardized["LONGITUDE_ROUND_5"] = df_standardized["LON_NAD83"].round(5)

    df_master["LATITUDE_ROUND_5"] = df_master["LAT_NAD83"].round(5)
    df_master["LONGITUDE_ROUND_5"] = df_master["LON_NAD83"].round(5)

    print("Executing Matching Algorithm...") #Executing Matching Algorithm...\n

    df_matched = algorithm.execute_matching_algorithm(
        df_standardized,
        df_master,
        match_scores_fields_path = os.path.join(workspace_directory, r"dev\matching\logic.json")
        )

    print("Generating Output Table...\n")

    df_scores_criteria = pd.DataFrame(
        [
            [1, "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "N/A", "N/A"],
            [2, "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Check for Mismatch", "Match", "Match", "N/A", "N/A"],
            [3, "Match", "Match", "Match", "Check for Mismatch", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "N/A", "N/A"],
            [4, "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Match", "N/A"],
            [5, "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Match", "N/A"],
            [6, "Match", "Match", "Match", "Check for Mismatch", "Match", "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "N/A"],
            [7, "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Match", "Match", "Match", "Match", "Match", "Match", "Check for Mismatch", "N/A"],
            [8, "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Match", "Match", "Match", "Match", "Match", "Match", "Match", "N/A"],
            [9, "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Match", "N/A"],
            [20, "Match", "Match", "Match", "Check for Mismatch", "Match", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Match", "Match", "Match", "Match", "Match"],
            [21, "Match", "Match", "Match", "Check for Mismatch", "Match", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Match", "Match", "Match", "Match", "Match"],
            [30, "Match", "Match", "Match", "Match", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch", "Check for Mismatch"]
            ],
        columns = ["Match_Score", "CO", "AB", "DIS", "FACID", "FNAME", "FSTREET", "FCITY", "FZIP", "FSIC","LAT", "LON", "Parcel", "FNAICS"]
        )

    df_final = algorithm.create_final_table(df, df_standardized, df_matched, df_scores_criteria, df_master)

    df_final.drop(columns=["UID"], inplace=True)

    df_final.to_excel(file_path_output.get(), index=False)

    wb = openpyxl.load_workbook(file_path_output.get())
    ws = wb['Sheet1']

    len_of_original_table = df.shape[1]

    for col in range(1, len_of_original_table):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")

    start_pos_of_standardized_table = len_of_original_table
    len_of_standardized_table = df_standardized.shape[1] - 1

    for col in range(start_pos_of_standardized_table, start_pos_of_standardized_table + len_of_standardized_table):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="FF7C80", end_color="FF7C80", fill_type="solid")

    start_pos_of_matched_table = len_of_standardized_table + start_pos_of_standardized_table
    len_of_matched_table = 2

    for col in range(start_pos_of_matched_table, start_pos_of_matched_table + len_of_matched_table):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="66FF33", end_color="66FF33", fill_type="solid")

    start_pos_of_scores_criteria_table = len_of_matched_table + start_pos_of_matched_table
    len_of_scores_criteria_table = df_scores_criteria.shape[1] - 1

    for col in range(start_pos_of_scores_criteria_table, start_pos_of_scores_criteria_table + len_of_scores_criteria_table):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")

    start_pos_of_master_table = len_of_scores_criteria_table + start_pos_of_scores_criteria_table
    len_of_master_table = df_master.shape[1] - 1

    for col in range(start_pos_of_master_table, start_pos_of_master_table + len_of_master_table):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="33CCFF", end_color="33CCFF", fill_type="solid")


    wb.save(file_path_output.get())

    print("Done!")

    df_summary = df_matched.groupby(['Match_Score']).size().reset_index('Match_Score')
    df_summary.columns = ['Match_Score', 'Match_Count']
    
    message_box_text = "Facility Matching Complete!\n\nMatch\tMatch\nScore\tCount\n\n"

    for index, row in df_summary.iterrows():
        message_box_text = "{0}{1}\t{2}\n".format(message_box_text, '{:,}'.format(int(row['Match_Score'])), '{:,}'.format(int(row['Match_Count'])))

    tk.messagebox.showinfo("FacFinder: CARB Facility Matching Tool", message_box_text) 


#################################################
########## CARB Facility Matching Code ##########
#################################################   


########## Main Application Window Frame ##########
    
## Create root for window
root = tk.Tk()
root.title("FacFinder")

## Set the height and width of the window
window_width = 800
window_height = 800
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width / 2) - (window_width / 2)
y = (screen_height / 2) - (window_height / 2)
root.geometry('%dx%d+%d+%d' % (window_width, window_height, x, y))


########## Title Frame ##########

## Create frame for title
title_frame = tk.Frame(root)
title_frame.pack(pady=10)

## Create a label for title
label = tk.Label(title_frame, text="FacFinder: CARB Facility Matching Tool")
label.pack(side='top', pady=10)
label.config(font=(text_font, 14, "bold"))


########## File Entry Frame ##########

## Create frame for file entry
file_selection_frame = tk.Frame(root)
file_selection_frame.pack(pady=10)

## Create a label for file selection
label = tk.Label(file_selection_frame, text="Select File")
label.grid(row=0, column=0, sticky='w')
label.config(font=(text_font, 10, "bold"))

## Load the folder icon image
folder_icon_path = os.path.join(workspace_directory, r"dev\icons\Custom-Icon-Design-Flatastic-1-Folder.512.png")
folder_icon = Image.open(folder_icon_path)
folder_icon = folder_icon.resize((14, 14))
folder_icon = ImageTk.PhotoImage(folder_icon)

## Create a button to open the file dialog
open_button = tk.Button(file_selection_frame, image=folder_icon, command=open_file_dialog, width=15, height=15)
open_button.grid(row=0, column=1, padx=5)

## Create a text box for displaying the selected file path
file_path_entry = tk.Entry(file_selection_frame, width=50, font=("Verdana", 11), state='readonly')
file_path_entry.grid(row=0, column=2, sticky='e')

## Create a label to select sheet
label = tk.Label(file_selection_frame, text="Select Sheet")
label.grid(row=1, column=0, sticky='w')
label.config(font=(text_font, 10, "bold"))

## Create dropdown to select sheet
options = [""]
selected_option = tk.StringVar(file_selection_frame)
selected_option.set(options[0])

## Update dropdown to contain list of sheets from excel file
dropdown = tk.OptionMenu(file_selection_frame, selected_option, *options)
dropdown.config(width=50)
dropdown.grid(row=1, column=2, sticky="ew", padx=0, pady=5) 

## Create load button to read in table
load_button = tk.Button(file_selection_frame, text="Load", command=load_button_actions)
load_button.config(font=(text_font, 10, "bold"))
load_button.grid(row=2, column=0, columnspan=3, pady=5, sticky='nsew')


########## Field Configuration Frame ##########

## Create frame for field configuration
field_selection_frame = tk.Frame(root)
field_selection_frame.pack(pady=10)

## Initialize dropdown options
field_options = [""]

## Create title for field configuration
selection_label = tk.Label(field_selection_frame, text="Field and Output Configuration")
selection_label.config(font=(text_font, 12, "bold"))

## CO -- Create field configuration
selected_option_CO = tk.StringVar(field_selection_frame)
selected_option_CO.set(options[0])
label_CO = tk.Label(field_selection_frame, text="CO")
label_CO.config(font=(text_font, 10, "bold"))
dropdown_CO = tk.OptionMenu(field_selection_frame, selected_option_CO, *field_options)
dropdown_CO.config(width=50)

## AB -- Create field configuration
selected_option_AB = tk.StringVar(field_selection_frame)
selected_option_AB.set(options[0])
label_AB = tk.Label(field_selection_frame, text="AB")
label_AB.config(font=(text_font, 10, "bold"))
dropdown_AB = tk.OptionMenu(field_selection_frame, selected_option_AB, *field_options)
dropdown_AB.config(width=50)

## DIS -- Create field configuration
selected_option_DIS = tk.StringVar(field_selection_frame)
selected_option_DIS.set(options[0])
label_DIS = tk.Label(field_selection_frame, text="DIS")
label_DIS.config(font=(text_font, 10, "bold"))
dropdown_DIS = tk.OptionMenu(field_selection_frame, selected_option_DIS, *field_options)
dropdown_DIS.config(width=50)

## FACID -- Create field configuration
selected_option_FACID = tk.StringVar(field_selection_frame)
selected_option_FACID.set(options[0])
label_FACID = tk.Label(field_selection_frame, text="Facility ID")
label_FACID.config(font=(text_font, 10, "bold"))
dropdown_FACID = tk.OptionMenu(field_selection_frame, selected_option_FACID, *field_options)
dropdown_FACID.config(width=50)

## FNAME -- Create field configuration
selected_option_FNAME = tk.StringVar(field_selection_frame)
selected_option_FNAME.set(options[0])
label_FNAME = tk.Label(field_selection_frame, text="Facility Name")
label_FNAME.config(font=(text_font, 10, "bold"))
dropdown_FNAME = tk.OptionMenu(field_selection_frame, selected_option_FNAME, *field_options)
dropdown_FNAME.config(width=50)

## FSTREET -- Create field configuration
selected_option_FSTREET = tk.StringVar(field_selection_frame)
selected_option_FSTREET.set(options[0])
label_FSTREET = tk.Label(field_selection_frame, text="Facility Street")
label_FSTREET.config(font=(text_font, 10, "bold"))
dropdown_FSTREET = tk.OptionMenu(field_selection_frame, selected_option_FSTREET, *field_options)
dropdown_FSTREET.config(width=50)

## FCITY -- Create field configuration
selected_option_FCITY = tk.StringVar(field_selection_frame)
selected_option_FCITY.set(options[0])
label_FCITY = tk.Label(field_selection_frame, text="Facility City")
label_FCITY.config(font=(text_font, 10, "bold"))
dropdown_FCITY = tk.OptionMenu(field_selection_frame, selected_option_FCITY, *field_options)
dropdown_FCITY.config(width=50)

## FZIP -- Create field configuration
selected_option_FZIP = tk.StringVar(field_selection_frame)
selected_option_FZIP.set(options[0])
label_FZIP = tk.Label(field_selection_frame, text="Facility ZIP")
label_FZIP.config(font=(text_font, 10, "bold"))
dropdown_FZIP = tk.OptionMenu(field_selection_frame, selected_option_FZIP, *field_options)
dropdown_FZIP.config(width=50)

## FSIC -- Create field configuration
selected_option_FSIC = tk.StringVar(field_selection_frame)
selected_option_FSIC.set(options[0])
label_FSIC = tk.Label(field_selection_frame, text="Facility SIC")
label_FSIC.config(font=(text_font, 10, "bold"))
dropdown_FSIC = tk.OptionMenu(field_selection_frame, selected_option_FSIC, *field_options)
dropdown_FSIC.config(width=50)

## FNAICS -- Create field configuration
selected_option_FNAICS = tk.StringVar(field_selection_frame)
selected_option_FNAICS.set(options[0])
label_FNAICS = tk.Label(field_selection_frame, text="Facility NAICS")
label_FNAICS.config(font=(text_font, 10, "bold"))
dropdown_FNAICS = tk.OptionMenu(field_selection_frame, selected_option_FNAICS, *field_options)
dropdown_FNAICS.config(width=50)

## Latitude -- Create field configuration
selected_option_Latitude = tk.StringVar(field_selection_frame)
selected_option_Latitude.set(options[0])
label_Latitude = tk.Label(field_selection_frame, text="Facility Latitude")
label_Latitude.config(font=(text_font, 10, "bold"))
dropdown_Latitude = tk.OptionMenu(field_selection_frame, selected_option_Latitude, *field_options)
dropdown_Latitude.config(width=50)

## Longitude -- Create field configuration
selected_option_Longitude = tk.StringVar(field_selection_frame)
selected_option_Longitude.set(options[0])
label_Longitude = tk.Label(field_selection_frame, text="Facility Longitude")
label_Longitude.config(font=(text_font, 10, "bold"))
dropdown_Longitude = tk.OptionMenu(field_selection_frame, selected_option_Longitude, *field_options)
dropdown_Longitude.config(width=50)


########## Create Output File Frame ##########

## Create frame for file output
file_output_frame = tk.Frame(root)
file_output_frame.pack(pady=10)

## Create title for file output
selection_output_label = tk.Label(file_output_frame, text="Output File Location")
selection_output_label.config(font=(text_font, 12, "bold"))

## Create a label for file output
label_output = tk.Label(file_output_frame, text="Output File")
label_output.config(font=(text_font, 10, "bold"))

## Create a button to open the file dialog
open_button_output = tk.Button(file_output_frame, image=folder_icon, command=select_output_file, width=15, height=15)

## Create a text box for displaying the selected file path
file_path_output = tk.Entry(file_output_frame, width=50, font=(text_font, 11), state='readonly')


########## Create Execute Button Frame ##########

## Create frame for execute button
execute_button_frame = tk.Frame(root)
execute_button_frame.pack(pady=5)

## Execute read in
execute_button = tk.Button(execute_button_frame, text="Execute", command=execute_facility_matching, padx=60, pady=10, fg="#000091")
execute_button.config(font=(text_font, 14, "bold"))


########## root.mainloop() ##########

root.mainloop()
